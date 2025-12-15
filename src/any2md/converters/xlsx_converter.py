"""XLSX to Markdown converter."""

import logging
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from doc2mkdocs.core.base_converter import BaseConverter, ConversionResult
from doc2mkdocs.core.config import ConversionConfig, ExcelMode

logger = logging.getLogger(__name__)


class XlsxConverter(BaseConverter):
    """Convert XLSX files to Markdown."""

    MAX_TABLE_ROWS = 1000
    MAX_TABLE_COLS = 50

    @property
    def supported_extensions(self) -> list[str]:
        """Get supported file extensions.

        Returns:
            List of supported extensions
        """
        return [".xlsx", ".xls"]

    def can_convert(self, file_path: Path) -> bool:
        """Check if this converter can handle the file.

        Args:
            file_path: Path to file

        Returns:
            True if file can be converted
        """
        return file_path.suffix.lower() in self.supported_extensions

    def convert(self, file_path: Path) -> ConversionResult:
        """Convert XLSX to Markdown.

        Args:
            file_path: Path to XLSX file

        Returns:
            Conversion result
        """
        result = ConversionResult(success=True)

        try:
            workbook = load_workbook(file_path, data_only=True)

            if self.config.excel_mode == ExcelMode.SINGLE_PAGE:
                # Combine all sheets into one markdown file
                markdown_parts = [f"# {file_path.stem}\n"]

                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    markdown_parts.append(f"\n## {sheet_name}\n")
                    sheet_md, warnings = self._convert_sheet(sheet)
                    markdown_parts.append(sheet_md)

                    for warning in warnings:
                        result.add_warning(f"Sheet '{sheet_name}': {warning}")

                result.markdown_content = "\n".join(markdown_parts)

            else:  # SHEET_PER_PAGE
                # For sheet-per-page mode, we'll just convert the first sheet
                # The CLI will handle creating multiple files
                if workbook.sheetnames:
                    sheet = workbook[workbook.sheetnames[0]]
                    markdown, warnings = self._convert_sheet(sheet)
                    result.markdown_content = markdown

                    for warning in warnings:
                        result.add_warning(warning)

                    # Store sheet names in metadata for CLI to handle
                    result.metadata["sheets"] = ",".join(workbook.sheetnames)
                    result.metadata["workbook_path"] = str(file_path)

            result.metadata["title"] = file_path.stem.replace("-", " ").title()

            workbook.close()

        except Exception as e:
            logger.error(f"Failed to convert {file_path}: {e}")
            result.add_error(f"Conversion failed: {str(e)}")

        return result

    def _convert_sheet(self, sheet: Worksheet) -> tuple[str, list[str]]:
        """Convert a single worksheet to Markdown.

        Args:
            sheet: Worksheet to convert

        Returns:
            Tuple of (markdown content, warnings)
        """
        warnings = []

        # Get dimensions
        max_row = sheet.max_row
        max_col = sheet.max_column

        # Check if sheet is too large
        if max_row > self.MAX_TABLE_ROWS or max_col > self.MAX_TABLE_COLS:
            warnings.append(
                f"Sheet is large ({max_row}x{max_col}), "
                f"truncating to {self.MAX_TABLE_ROWS}x{self.MAX_TABLE_COLS}"
            )
            max_row = min(max_row, self.MAX_TABLE_ROWS)
            max_col = min(max_col, self.MAX_TABLE_COLS)

        # Check if sheet is empty
        if max_row == 0 or max_col == 0:
            return "_Empty sheet_\n", warnings

        # Try to convert to markdown table
        try:
            markdown = self._sheet_to_markdown_table(sheet, max_row, max_col)
        except Exception as e:
            warnings.append(f"Failed to create Markdown table, using HTML fallback: {e}")
            markdown = self._sheet_to_html_table(sheet, max_row, max_col)

        return markdown, warnings

    def _sheet_to_markdown_table(
        self, sheet: Worksheet, max_row: int, max_col: int
    ) -> str:
        """Convert sheet to Markdown table.

        Args:
            sheet: Worksheet
            max_row: Maximum row to include
            max_col: Maximum column to include

        Returns:
            Markdown table
        """
        lines = []

        # Get all rows
        rows = []
        for row_idx in range(1, max_row + 1):
            row_data = []
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row_idx, col_idx)
                value = cell.value
                # Convert to string and escape pipes
                cell_text = str(value) if value is not None else ""
                cell_text = cell_text.replace("|", "\\|").replace("\n", " ")
                row_data.append(cell_text)
            rows.append(row_data)

        if not rows:
            return "_Empty sheet_\n"

        # First row as header
        header = rows[0]
        lines.append("| " + " | ".join(header) + " |")

        # Separator
        lines.append("| " + " | ".join(["---"] * len(header)) + " |")

        # Data rows
        for row in rows[1:]:
            # Pad row to match header length
            while len(row) < len(header):
                row.append("")
            lines.append("| " + " | ".join(row[:len(header)]) + " |")

        return "\n".join(lines) + "\n"

    def _sheet_to_html_table(
        self, sheet: Worksheet, max_row: int, max_col: int
    ) -> str:
        """Convert sheet to HTML table (fallback).

        Args:
            sheet: Worksheet
            max_row: Maximum row to include
            max_col: Maximum column to include

        Returns:
            HTML table
        """
        lines = ["<table>"]

        for row_idx in range(1, max_row + 1):
            lines.append("  <tr>")
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row_idx, col_idx)
                value = cell.value
                cell_text = str(value) if value is not None else ""
                # Escape HTML
                cell_text = (
                    cell_text.replace("&", "&amp;")
                    .replace("<", "&lt;")
                    .replace(">", "&gt;")
                )
                tag = "th" if row_idx == 1 else "td"
                lines.append(f"    <{tag}>{cell_text}</{tag}>")
            lines.append("  </tr>")

        lines.append("</table>")
        return "\n".join(lines) + "\n"

